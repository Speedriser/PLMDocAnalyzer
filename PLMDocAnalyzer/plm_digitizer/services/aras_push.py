"""
PLM Digitizer - Aras PLM Push Service
Reads output file and pushes records to Aras Innovator via REST API.
"""
import base64
import csv
import json
import logging
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Generator, List, Optional, Tuple

import httpx

logger = logging.getLogger(__name__)

BATCH_SIZE = 50


def read_output_file(file_path: str) -> Tuple[List[str], List[Dict]]:
    """
    Read output file (Excel or CSV) and return (headers, rows).
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Output file not found: {file_path}")

    ext = path.suffix.lower()

    if ext in (".xlsx", ".xls"):
        return _read_excel(file_path)
    elif ext == ".csv":
        return _read_csv(file_path)
    else:
        raise ValueError(f"Unsupported output format: {ext}")


def _read_excel(file_path: str) -> Tuple[List[str], List[Dict]]:
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    # Try to find "Results" sheet first
    sheet_name = "Results" if "Results" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data_rows = []
    for row in rows[1:]:
        row_dict = {
            headers[i]: str(v).strip() if v is not None else ""
            for i, v in enumerate(row)
            if i < len(headers)
        }
        # Skip empty rows
        if any(v for v in row_dict.values()):
            data_rows.append(row_dict)

    wb.close()
    return headers, data_rows


def _read_csv(file_path: str) -> Tuple[List[str], List[Dict]]:
    rows = []
    headers = []
    with open(file_path, "r", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.DictReader(f)
        headers = list(reader.fieldnames or [])
        for row in reader:
            rows.append(dict(row))
    return headers, rows


def build_aml_item(item_type: str, row: Dict, field_mapping: Dict[str, str]) -> str:
    """
    Build AML XML for a single record.
    """
    item = ET.Element("Item")
    item.set("type", item_type)
    item.set("action", "add")

    for output_col, aras_prop in field_mapping.items():
        value = row.get(output_col, "").strip()
        if value and aras_prop:
            prop = ET.SubElement(item, aras_prop)
            prop.text = value

    return ET.tostring(item, encoding="unicode")


def build_aml_batch(items_aml: List[str]) -> str:
    """Wrap multiple item AMLs in AML envelope."""
    aml = ET.Element("AML")
    for item_str in items_aml:
        item_el = ET.fromstring(item_str)
        aml.append(item_el)
    return '<?xml version="1.0" encoding="utf-8"?>' + ET.tostring(aml, encoding="unicode")


def test_aras_connection(
    server_url: str,
    database_name: Optional[str],
    username: Optional[str],
    password: Optional[str],
) -> Tuple[bool, str]:
    """
    Test connectivity to Aras Innovator server.
    Returns (success, message).
    """
    try:
        base_url = server_url.rstrip("/")
        # Try to reach the server
        with httpx.Client(timeout=10.0, verify=False) as client:
            # Try a basic health endpoint
            try:
                resp = client.get(f"{base_url}/Server/odata/", timeout=10.0)
                if resp.status_code in (200, 401, 403):
                    return True, f"Connected to Aras server (HTTP {resp.status_code})"
            except Exception:
                pass

            # Try root URL
            resp = client.get(base_url, timeout=10.0)
            if resp.status_code < 500:
                return True, f"Server reachable (HTTP {resp.status_code})"
            return False, f"Server returned HTTP {resp.status_code}"

    except httpx.ConnectError:
        return False, f"Cannot connect to server: {server_url}"
    except httpx.TimeoutException:
        return False, f"Connection timeout: {server_url}"
    except Exception as e:
        return False, f"Connection error: {str(e)}"


class ArasPusher:
    """Handles pushing records to Aras Innovator."""

    def __init__(
        self,
        server_url: str,
        database_name: Optional[str],
        username: Optional[str],
        password: Optional[str],
        item_type: str,
    ):
        self.server_url = server_url.rstrip("/")
        self.database_name = database_name or ""
        self.username = username or ""
        self.password = password or ""
        self.item_type = item_type

        # Build auth header
        if username and password:
            credentials = base64.b64encode(f"{username}:{password}".encode()).decode()
            self.auth_header = f"Basic {credentials}"
        else:
            self.auth_header = None

    def _get_headers(self) -> Dict[str, str]:
        headers = {
            "Content-Type": "text/xml; charset=utf-8",
            "Accept": "application/json",
            "AUTHUSER": self.username,
            "AUTHPASSWORD": self.password,
            "DATABASE": self.database_name,
        }
        if self.auth_header:
            headers["Authorization"] = self.auth_header
        return headers

    def push_record(self, row: Dict, field_mapping: Dict[str, str]) -> Tuple[bool, str]:
        """Push a single record to Aras."""
        try:
            # Build JSON body for OData REST
            body = {}
            for output_col, aras_prop in field_mapping.items():
                value = row.get(output_col, "").strip()
                if value and aras_prop:
                    body[aras_prop] = value

            if not body:
                return False, "No fields to push"

            endpoint = f"{self.server_url}/Server/odata/{self.item_type}"

            with httpx.Client(timeout=30.0, verify=False) as client:
                resp = client.post(
                    endpoint,
                    json=body,
                    headers={
                        "Content-Type": "application/json",
                        "Authorization": self.auth_header or "",
                        "DATABASE": self.database_name,
                    },
                )

                if resp.status_code in (200, 201):
                    return True, f"Created record (HTTP {resp.status_code})"
                else:
                    return False, f"HTTP {resp.status_code}: {resp.text[:200]}"

        except Exception as e:
            return False, str(e)

    def push_batch_aml(
        self, rows: List[Dict], field_mapping: Dict[str, str]
    ) -> Tuple[int, int, List[str]]:
        """
        Push a batch via AML.
        Returns (passed, failed, error_messages).
        """
        passed = 0
        failed = 0
        errors = []

        items_aml = []
        for row in rows:
            aml = build_aml_item(self.item_type, row, field_mapping)
            items_aml.append(aml)

        # For now, push individually with OData REST
        for i, row in enumerate(rows):
            ok, msg = self.push_record(row, field_mapping)
            if ok:
                passed += 1
            else:
                failed += 1
                errors.append(f"Row {i}: {msg}")

        return passed, failed, errors


def push_to_aras(
    output_file: str,
    connection_config: Dict,
    field_mapping: Dict[str, str],
    progress_callback=None,
    retry_failed_ids: Optional[List[int]] = None,
) -> Dict:
    """
    Main push function. Reads output file and pushes all records to Aras.

    progress_callback(pushed, total, failed) called periodically.
    Returns summary dict.
    """
    headers, rows = read_output_file(output_file)

    if not rows:
        return {"success": False, "error": "No data rows found in output file"}

    if retry_failed_ids:
        rows = [r for i, r in enumerate(rows) if i in retry_failed_ids]

    pusher = ArasPusher(
        server_url=connection_config["server_url"],
        database_name=connection_config.get("database_name"),
        username=connection_config.get("username"),
        password=connection_config.get("password"),
        item_type=connection_config.get("item_type", "Part"),
    )

    total = len(rows)
    pushed = 0
    failed = 0
    all_errors = []

    # Process in batches
    batch_size = BATCH_SIZE
    for batch_start in range(0, total, batch_size):
        batch = rows[batch_start : batch_start + batch_size]
        p, f, errs = pusher.push_batch_aml(batch, field_mapping)
        pushed += p
        failed += f
        all_errors.extend(errs)

        if progress_callback:
            try:
                progress_callback(pushed + failed, total, failed)
            except Exception:
                pass

    return {
        "success": True,
        "total": total,
        "pushed": pushed,
        "failed": failed,
        "errors": all_errors[:100],  # Limit error list
        "completed_at": datetime.utcnow().isoformat(),
    }
