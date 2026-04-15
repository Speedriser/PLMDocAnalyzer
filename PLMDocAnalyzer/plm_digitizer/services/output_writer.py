"""
PLM Digitizer - Output File Writer
Writes results to Excel (with color coding) or CSV as batches complete.

Column order:
  1. File Path       — full path to the source file
  2. File Name       — basename of the source file
  3. Confidence Score — numeric score 0.00-1.00
  4..N  Extracted fields (as configured by the user)
"""
import csv
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

# Color constants for Excel (ARGB)
GREEN_FILL  = "FF90EE90"   # confidence > 0.8
YELLOW_FILL = "FFFFD700"   # confidence 0.5-0.8
RED_FILL    = "FFFF6B6B"   # confidence < 0.5
HEADER_FILL = "FF2D3748"   # Dark header
HEADER_FONT_COLOR = "FFFFFFFF"

# Fixed leading columns (always first three)
LEAD_COLUMNS = ["File Path", "File Name", "Confidence Score"]


class ExcelWriter:
    """Streaming Excel writer using openpyxl write-only mode."""

    def __init__(
        self,
        output_path: str,
        fields: List[str],
        confidence_threshold: float = 0.7,
    ):
        from openpyxl import Workbook

        self.output_path = output_path
        self.fields = fields
        self.confidence_threshold = confidence_threshold
        self.row_count = 0
        self.passed = 0
        self.failed = 0

        # Create workbook in write-only mode
        self.wb = Workbook(write_only=True)
        self.ws = self.wb.create_sheet("Results")
        self.summary_ws = self.wb.create_sheet("Summary")

        self._write_header()

    def _get_header_row(self) -> List[str]:
        # File Path | File Name | Confidence Score | <user fields...>
        return LEAD_COLUMNS + self.fields

    def _write_header(self):
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.cell import WriteOnlyCell

        cells = []
        for col_name in self._get_header_row():
            cell = WriteOnlyCell(self.ws, value=col_name)
            cell.font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
            cell.fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cells.append(cell)
        self.ws.append(cells)

    def write_row(
        self,
        file_path: str,
        file_type: str,
        extracted_data: Optional[Dict],
        confidence: float,
        status: str,
        processing_time_ms: int = 0,
        extraction_method: str = "",
    ):
        from openpyxl.styles import PatternFill
        from openpyxl.cell import WriteOnlyCell

        self.row_count += 1
        if status == "passed":
            self.passed += 1
        else:
            self.failed += 1

        # Row fill color based on confidence
        if confidence >= 0.8:
            fill_color = GREEN_FILL
        elif confidence >= 0.5:
            fill_color = YELLOW_FILL
        else:
            fill_color = RED_FILL
        fill = PatternFill(fill_type="solid", fgColor=fill_color)

        data = extracted_data or {}

        # Build row: File Path, File Name, Confidence Score, then extracted fields
        row_values = [
            file_path,
            os.path.basename(file_path),
            f"{confidence:.2f}" if confidence is not None else "",
        ]

        for field in self.fields:
            val = data.get(field)
            if val is None:
                val = data.get(field.lower())
            if val is None:
                val = data.get(field.replace(" ", "_").lower())
            row_values.append(str(val) if val is not None else "")

        cells = []
        for val in row_values:
            cell = WriteOnlyCell(self.ws, value=val)
            cell.fill = fill
            cells.append(cell)
        self.ws.append(cells)

    def finalize(self, run_stats: Optional[Dict] = None):
        """Write summary sheet and save workbook."""
        from openpyxl.styles import Font, PatternFill
        from openpyxl.cell import WriteOnlyCell

        stats = run_stats or {}

        # Summary header
        h_cells = []
        for h in ["Metric", "Value"]:
            cell = WriteOnlyCell(self.summary_ws, value=h)
            cell.font = Font(bold=True, color=HEADER_FONT_COLOR)
            cell.fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)
            h_cells.append(cell)
        self.summary_ws.append(h_cells)

        summary_rows = [
            ("Run Name",       stats.get("name", "")),
            ("Run ID",         stats.get("id", "")),
            ("Total Files",    stats.get("total_files", self.row_count)),
            ("Passed Records", self.passed),
            ("Failed Records", self.failed),
            ("Pass Rate",      f"{100 * self.passed // max(self.row_count, 1)}%"),
            ("LLM Model",      stats.get("llm_model", "")),
            ("Generated At",   datetime.utcnow().isoformat()),
        ]
        for metric, value in summary_rows:
            self.summary_ws.append([metric, str(value)])

        Path(self.output_path).parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(self.output_path)
        logger.info(f"Excel output saved to {self.output_path}")
        return self.output_path


class CSVWriter:
    """Streaming CSV writer."""

    def __init__(self, output_path: str, fields: List[str]):
        self.output_path = output_path
        self.fields = fields
        self.row_count = 0
        self.passed = 0
        self.failed = 0

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        self._file = open(output_path, "w", newline="", encoding="utf-8-sig")
        self._writer = csv.writer(self._file)

        # Header: File Path | File Name | Confidence Score | <user fields...>
        self._writer.writerow(LEAD_COLUMNS + fields)

    def write_row(
        self,
        file_path: str,
        file_type: str,
        extracted_data: Optional[Dict],
        confidence: float,
        status: str,
        processing_time_ms: int = 0,
        extraction_method: str = "",
    ):
        self.row_count += 1
        if status == "passed":
            self.passed += 1
        else:
            self.failed += 1

        data = extracted_data or {}

        row_values = [
            file_path,
            os.path.basename(file_path),
            f"{confidence:.2f}" if confidence is not None else "",
        ]

        for field in self.fields:
            val = data.get(field)
            if val is None:
                val = data.get(field.lower())
            if val is None:
                val = data.get(field.replace(" ", "_").lower())
            row_values.append(str(val) if val is not None else "")

        self._writer.writerow(row_values)

    def finalize(self, run_stats: Optional[Dict] = None):
        """Flush and close the CSV file."""
        self._file.flush()
        self._file.close()
        logger.info(f"CSV output saved to {self.output_path}")
        return self.output_path


def create_writer(
    output_path: str,
    output_format: str,
    fields: List[str],
    confidence_threshold: float = 0.7,
) -> Any:
    """Factory function to create the appropriate writer."""
    if output_format.lower() == "excel":
        return ExcelWriter(output_path, fields, confidence_threshold)
    else:
        return CSVWriter(output_path, fields)


def generate_output_path(
    run_name: str,
    output_format: str,
    base_dir: Optional[str] = None,
) -> str:
    """Generate a default output file path."""
    from config import DATA_DIR

    if not base_dir:
        base_dir = str(DATA_DIR / "outputs")

    Path(base_dir).mkdir(parents=True, exist_ok=True)

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in run_name)
    ext = "xlsx" if output_format.lower() == "excel" else "csv"

    return str(Path(base_dir) / f"{safe_name}_{timestamp}.{ext}")
